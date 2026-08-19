# reportes/views.py
#
# FASE 1 — CAMBIO 1: Reporte de Citas (filtros + datos en pantalla). Ya probado.
# FASE 1 — CAMBIO 2: Exportación real a PDF del Reporte de Citas, con reportlab.
#
# La lógica de filtrado se extrajo a _filtrar_citas() para que el JSON (pantalla)
# y el PDF usen exactamente los mismos filtros y no haya inconsistencias entre
# lo que el usuario ve en pantalla y lo que descarga.

from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse

from citas.models import Cita
from usuarios.permisos import solo_admin


def _filtrar_citas(request):
    """
    Aplica los filtros de fecha_desde, fecha_hasta, asesor y estado sobre
    Cita.objects, leyendo los parámetros de request.GET.
    Devuelve (queryset, lista_de_errores, dict_de_filtros_aplicados).
    """
    qs = Cita.objects.select_related('asesor').all()

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    asesor_id   = request.GET.get('asesor', '').strip()
    estado      = request.GET.get('estado', '').strip()

    errores = []
    filtros_aplicados = {
        'fecha_desde': None,
        'fecha_hasta': None,
        'asesor':      None,
        'estado':      None,
    }

    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            qs = qs.filter(fecha__gte=fd)
            filtros_aplicados['fecha_desde'] = fd.strftime('%d/%m/%Y')
        except ValueError:
            errores.append('Fecha desde inválida.')

    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            qs = qs.filter(fecha__lte=fh)
            filtros_aplicados['fecha_hasta'] = fh.strftime('%d/%m/%Y')
        except ValueError:
            errores.append('Fecha hasta inválida.')

    if asesor_id:
        try:
            asesor_obj = User.objects.filter(id=int(asesor_id)).first()
            qs = qs.filter(asesor_id=int(asesor_id))
            if asesor_obj:
                filtros_aplicados['asesor'] = asesor_obj.get_full_name() or asesor_obj.username
        except ValueError:
            errores.append('Asesor inválido.')

    if estado:
        if estado in dict(Cita.ESTADO_CHOICES):
            qs = qs.filter(estado=estado)
            filtros_aplicados['estado'] = dict(Cita.ESTADO_CHOICES)[estado]
        else:
            errores.append('Estado inválido.')

    return qs, errores, filtros_aplicados


@login_required
@solo_admin
def reporte_citas_ajax(request):
    """
    Devuelve estadísticas y listado de citas filtrado por:
      - fecha_desde (YYYY-MM-DD)
      - fecha_hasta (YYYY-MM-DD)
      - asesor (id de User)
      - estado (pendiente | confirmada | completada | cancelada)

    Todos los filtros son opcionales.
    """
    qs, errores, _ = _filtrar_citas(request)

    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total       = qs.count()
    pendientes  = qs.filter(estado='pendiente').count()
    confirmadas = qs.filter(estado='confirmada').count()
    atendidas   = qs.filter(estado='completada').count()
    canceladas  = qs.filter(estado='cancelada').count()

    # Citas por asesor (solo asesores con al menos 1 cita en el filtro actual)
    por_asesor = {}
    for c in qs:
        if c.asesor:
            nombre = c.asesor.get_full_name() or c.asesor.username
        else:
            nombre = 'Sin asignar'
        por_asesor[nombre] = por_asesor.get(nombre, 0) + 1
    citas_por_asesor = [
        {'asesor': nombre, 'total': cnt}
        for nombre, cnt in sorted(por_asesor.items(), key=lambda x: -x[1])
    ]

    # Citas por periodo (agrupadas por fecha exacta, para graficar tendencia)
    por_fecha = {}
    for c in qs:
        clave = c.fecha.strftime('%Y-%m-%d')
        por_fecha[clave] = por_fecha.get(clave, 0) + 1
    citas_por_periodo = [
        {'fecha': fecha, 'total': cnt}
        for fecha, cnt in sorted(por_fecha.items())
    ]

    # Listado detallado (máx. 200 filas para no sobrecargar la respuesta;
    # el total real ya viene arriba en 'total')
    listado = [
        {
            'id':      c.id,
            'cliente': f'{c.nombre_cliente} {c.apellidos_cliente}',
            'fecha':   c.fecha.strftime('%d/%m/%Y'),
            'hora':    c.hora.strftime('%H:%M'),
            'motivo':  c.get_motivo_display(),
            'estado':  c.get_estado_display(),
            'asesor':  (c.asesor.get_full_name() or c.asesor.username) if c.asesor else 'Sin asignar',
        }
        for c in qs.order_by('-fecha', '-hora')[:200]
    ]

    return JsonResponse({
        'ok': True,
        'resumen': {
            'total':       total,
            'pendientes':  pendientes,
            'confirmadas': confirmadas,
            'atendidas':   atendidas,
            'canceladas':  canceladas,
        },
        'citas_por_asesor':   citas_por_asesor,
        'citas_por_periodo':  citas_por_periodo,
        'listado':            listado,
    })


@login_required
@solo_admin
def asesores_ajax(request):
    """Lista de asesores/admins activos, para poblar el filtro de 'Asesor' en Reportes."""
    usuarios = (
        User.objects
        .filter(is_active=True, perfil__isnull=False)
        .select_related('perfil')
        .order_by('first_name', 'username')
    )
    data = [
        {
            'id':     u.id,
            'nombre': u.get_full_name() or u.username,
            'rol':    u.perfil.get_rol_display(),
        }
        for u in usuarios
    ]
    return JsonResponse({'ok': True, 'asesores': data})


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS GENÉRICOS DE EXPORTACIÓN (usados por todos los reportes nuevos)
# ═══════════════════════════════════════════════════════════════════════════

def _pdf_generico(titulo, subtitulo, resumen_headers, resumen_valores,
                   tabla_headers, tabla_filas, col_widths, nombre_archivo):
    """
    Construye un PDF apaisado con: título, subtítulo (filtros aplicados +
    fecha de generación), una tabla de resumen (KPIs) y una tabla detallada.
    Reutilizado por Prospectos, Postulantes, Vacantes, Servicios y Resumen
    General para no duplicar el boilerplate de reportlab que ya se probó
    en exportar_citas_pdf().
    """
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'TituloVR', parent=styles['Heading1'],
        textColor=colors.HexColor('#1560BD'), fontSize=18, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'SubVR', parent=styles['Normal'],
        textColor=colors.HexColor('#64748b'), fontSize=9, spaceAfter=12,
    )

    elementos = [Paragraph(titulo, titulo_style), Paragraph(subtitulo, sub_style)]

    if resumen_headers:
        resumen_data = [resumen_headers, [str(v) for v in resumen_valores]]
        ancho_col = (25.5 * cm) / len(resumen_headers)
        tabla_resumen = Table(resumen_data, colWidths=[ancho_col] * len(resumen_headers))
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1560BD')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 10),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f1f5f9')),
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 18))

    filas = [tabla_headers] + tabla_filas
    if len(filas) == 1:
        filas.append(['Sin datos para los filtros seleccionados.'] + [''] * (len(tabla_headers) - 1))

    if not col_widths:
        col_widths = [(25.5 * cm) / len(tabla_headers)] * len(tabla_headers)

    tabla_listado = Table(filas, colWidths=col_widths, repeatRows=1)
    tabla_listado.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f1e35')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_listado)

    doc.build(elementos)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


def _excel_generico(hoja_titulo, resumen_headers, resumen_valores,
                     tabla_headers, tabla_filas, nombre_archivo):
    """
    Construye un archivo .xlsx real con openpyxl: una fila de KPIs (si se
    proveen) y la tabla detallada, con encabezados estilizados y columnas
    autoajustadas. Reutilizado por todos los reportes nuevos.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = hoja_titulo[:31] or 'Reporte'

    header_fill = PatternFill(start_color='0F1E35', end_color='0F1E35', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    kpi_fill    = PatternFill(start_color='1560BD', end_color='1560BD', fill_type='solid')
    kpi_font    = Font(color='FFFFFF', bold=True)

    fila_actual = 1

    if resumen_headers:
        for col, texto in enumerate(resumen_headers, start=1):
            celda = ws.cell(row=fila_actual, column=col, value=texto)
            celda.fill = kpi_fill
            celda.font = kpi_font
            celda.alignment = Alignment(horizontal='center')
        fila_actual += 1
        for col, valor in enumerate(resumen_valores, start=1):
            ws.cell(row=fila_actual, column=col, value=valor).alignment = Alignment(horizontal='center')
        fila_actual += 2  # fila en blanco de separación

    for col, texto in enumerate(tabla_headers, start=1):
        celda = ws.cell(row=fila_actual, column=col, value=texto)
        celda.fill = header_fill
        celda.font = header_font
    fila_actual += 1

    for fila in tabla_filas:
        for col, valor in enumerate(fila, start=1):
            ws.cell(row=fila_actual, column=col, value=valor)
        fila_actual += 1

    # Autoajuste simple de ancho de columna
    n_cols = max(len(resumen_headers or []), len(tabla_headers))
    for col in range(1, n_cols + 1):
        letra = get_column_letter(col)
        max_len = max(
            [len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, fila_actual)] or [10]
        )
        ws.column_dimensions[letra].width = min(max_len + 3, 45)

    buffer = BytesIO()
    wb.save(buffer)
    contenido = buffer.getvalue()
    buffer.close()

    response = HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
@solo_admin
def exportar_citas_pdf(request):
    """
    FASE 1 — CAMBIO 2.
    Genera un PDF real (reportlab) del Reporte de Citas, con los mismos
    filtros que la vista en pantalla. Sin límite de 200 filas: el PDF
    incluye TODAS las citas que cumplan el filtro.
    """
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    qs, errores, filtros_aplicados = _filtrar_citas(request)

    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    citas = list(qs.order_by('-fecha', '-hora'))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'TituloVR', parent=styles['Heading1'],
        textColor=colors.HexColor('#1560BD'), fontSize=18, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'SubVR', parent=styles['Normal'],
        textColor=colors.HexColor('#64748b'), fontSize=9, spaceAfter=12,
    )

    elementos = []
    elementos.append(Paragraph('Grupo V&amp;R Consultores — Reporte de Citas', titulo_style))

    filtros_texto = []
    if filtros_aplicados['fecha_desde']:
        filtros_texto.append(f"Desde: {filtros_aplicados['fecha_desde']}")
    if filtros_aplicados['fecha_hasta']:
        filtros_texto.append(f"Hasta: {filtros_aplicados['fecha_hasta']}")
    if filtros_aplicados['asesor']:
        filtros_texto.append(f"Asesor: {filtros_aplicados['asesor']}")
    if filtros_aplicados['estado']:
        filtros_texto.append(f"Estado: {filtros_aplicados['estado']}")
    generado = datetime.now().strftime('%d/%m/%Y %H:%M')
    linea_filtros = (' | '.join(filtros_texto) if filtros_texto else 'Sin filtros aplicados') + f' — Generado: {generado}'
    elementos.append(Paragraph(linea_filtros, sub_style))

    # ── Resumen ──────────────────────────────────────────────────────
    total       = len(citas)
    pendientes  = sum(1 for c in citas if c.estado == 'pendiente')
    confirmadas = sum(1 for c in citas if c.estado == 'confirmada')
    atendidas   = sum(1 for c in citas if c.estado == 'completada')
    canceladas  = sum(1 for c in citas if c.estado == 'cancelada')

    resumen_data = [
        ['Total', 'Pendientes', 'Confirmadas', 'Atendidas', 'Canceladas'],
        [str(total), str(pendientes), str(confirmadas), str(atendidas), str(canceladas)],
    ]
    tabla_resumen = Table(resumen_data, colWidths=[4.6 * cm] * 5)
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1560BD')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f1f5f9')),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 18))

    # ── Listado detallado ───────────────────────────────────────────
    encabezados = ['Cliente', 'Fecha', 'Hora', 'Motivo', 'Estado', 'Asesor']
    filas = [encabezados]
    for c in citas:
        filas.append([
            f'{c.nombre_cliente} {c.apellidos_cliente}',
            c.fecha.strftime('%d/%m/%Y'),
            c.hora.strftime('%H:%M'),
            c.get_motivo_display(),
            c.get_estado_display(),
            (c.asesor.get_full_name() or c.asesor.username) if c.asesor else 'Sin asignar',
        ])

    if len(filas) == 1:
        filas.append(['Sin citas para los filtros seleccionados.', '', '', '', '', ''])

    tabla_listado = Table(filas, colWidths=[5.5 * cm, 2.8 * cm, 2 * cm, 4.5 * cm, 3 * cm, 4.6 * cm], repeatRows=1)
    tabla_listado.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f1e35')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_listado)

    doc.build(elementos)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    nombre_archivo = f'reporte_citas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

# ═══════════════════════════════════════════════════════════════════════════
# FASE 1 — MÓDULO GENERAL DE REPORTES (ampliación completa)
#
# Prospectos, Postulantes, Vacantes, Servicios, Resumen General y Actividad
# General. Mismo criterio que Citas: una función _filtrar_x() compartida
# entre pantalla (AJAX/JSON), PDF y Excel para que los tres canales muestren
# siempre exactamente los mismos datos.
# ═══════════════════════════════════════════════════════════════════════════

from datetime import datetime as _dt, time as _time, timedelta as _timedelta
from django.utils import timezone as _timezone


def _rango_periodo(periodo):
    """Convierte 'hoy'|'semana'|'mes'|'año'|'' en un datetime de inicio (o None si es 'todo')."""
    hoy = _timezone.now().date()
    if periodo == 'hoy':
        inicio = hoy
    elif periodo == 'semana':
        inicio = hoy - _timedelta(days=hoy.weekday())
    elif periodo == 'mes':
        inicio = hoy.replace(day=1)
    elif periodo == 'año':
        inicio = hoy.replace(month=1, day=1)
    else:
        return None
    return _timezone.make_aware(_dt.combine(inicio, _time.min))


# ── PROSPECTOS ──────────────────────────────────────────────────────────────

def _filtrar_prospectos(request):
    from prospectos.models import Prospecto

    qs = Prospecto.objects.select_related('asesor_asignado').all()

    periodo   = request.GET.get('periodo', '').strip()
    asesor_id = request.GET.get('asesor', '').strip()
    estado    = request.GET.get('estado', '').strip()

    errores = []
    filtros_aplicados = {'periodo': None, 'asesor': None, 'estado': None}

    inicio = _rango_periodo(periodo)
    if inicio:
        qs = qs.filter(fecha__gte=inicio)
        filtros_aplicados['periodo'] = {'hoy': 'Hoy', 'semana': 'Esta semana',
                                         'mes': 'Este mes', 'año': 'Este año'}.get(periodo)

    if asesor_id:
        try:
            asesor_obj = User.objects.filter(id=int(asesor_id)).first()
            qs = qs.filter(asesor_asignado_id=int(asesor_id))
            if asesor_obj:
                filtros_aplicados['asesor'] = asesor_obj.get_full_name() or asesor_obj.username
        except ValueError:
            errores.append('Asesor inválido.')

    if estado:
        from prospectos.models import Prospecto as P
        if estado in dict(P.ESTADO_CHOICES):
            qs = qs.filter(estado=estado)
            filtros_aplicados['estado'] = dict(P.ESTADO_CHOICES)[estado]
        else:
            errores.append('Estado inválido.')

    return qs, errores, filtros_aplicados


def _resumen_prospectos(qs):
    from prospectos.models import Prospecto

    total = qs.count()
    por_estado = {
        clave: qs.filter(estado=clave).count()
        for clave, _ in Prospecto.ESTADO_CHOICES
    }
    por_asesor_map = {}
    for p in qs:
        nombre = (p.asesor_asignado.get_full_name() or p.asesor_asignado.username) if p.asesor_asignado else 'Sin asignar'
        por_asesor_map[nombre] = por_asesor_map.get(nombre, 0) + 1
    por_asesor = [{'asesor': n, 'total': c} for n, c in sorted(por_asesor_map.items(), key=lambda x: -x[1])]

    return total, por_estado, por_asesor


@login_required
@solo_admin
def reporte_prospectos_ajax(request):
    from prospectos.models import Prospecto

    qs, errores, _ = _filtrar_prospectos(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, por_asesor = _resumen_prospectos(qs)
    estado_labels = dict(Prospecto.ESTADO_CHOICES)

    listado = [
        {
            'id':      p.id,
            'nombre':  p.nombre,
            'correo':  p.correo,
            'telefono': p.telefono,
            'interes': p.get_interes_display(),
            'estado':  p.get_estado_display(),
            'asesor':  (p.asesor_asignado.get_full_name() or p.asesor_asignado.username) if p.asesor_asignado else 'Sin asignar',
            'fecha':   p.fecha.strftime('%d/%m/%Y'),
        }
        for p in qs.order_by('-fecha')[:200]
    ]

    return JsonResponse({
        'ok': True,
        'resumen': {
            'total':       total,
            'nuevo':       por_estado.get('nuevo', 0),
            'contactado':  por_estado.get('contactado', 0),
            'convertido':  por_estado.get('convertido', 0),
            'descartado':  por_estado.get('descartado', 0),
        },
        'por_estado':  [{'estado': estado_labels[k], 'total': v} for k, v in por_estado.items()],
        'por_asesor':  por_asesor,
        'listado':     listado,
    })


@login_required
@solo_admin
def exportar_prospectos_pdf(request):
    qs, errores, filtros_aplicados = _filtrar_prospectos(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, _ = _resumen_prospectos(qs)

    filtros_texto = [v for k, v in filtros_aplicados.items() if v]
    generado = _dt.now().strftime('%d/%m/%Y %H:%M')
    subtitulo = (' | '.join(f'{k.capitalize()}: {v}' for k, v in filtros_aplicados.items() if v)
                 if filtros_texto else 'Sin filtros aplicados') + f' — Generado: {generado}'

    tabla_filas = [
        [p.nombre, p.correo, p.telefono, p.get_interes_display(), p.get_estado_display(),
         (p.asesor_asignado.get_full_name() or p.asesor_asignado.username) if p.asesor_asignado else 'Sin asignar',
         p.fecha.strftime('%d/%m/%Y')]
        for p in qs.order_by('-fecha')
    ]

    return _pdf_generico(
        titulo='Grupo V&amp;R Consultores — Reporte de Prospectos',
        subtitulo=subtitulo,
        resumen_headers=['Total', 'Nuevos', 'Contactados', 'Convertidos', 'Descartados'],
        resumen_valores=[total, por_estado.get('nuevo', 0), por_estado.get('contactado', 0),
                          por_estado.get('convertido', 0), por_estado.get('descartado', 0)],
        tabla_headers=['Nombre', 'Correo', 'Teléfono', 'Interés', 'Estado', 'Asesor', 'Fecha'],
        tabla_filas=tabla_filas,
        col_widths=None,
        nombre_archivo=f'reporte_prospectos_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf',
    )


@login_required
@solo_admin
def exportar_prospectos_excel(request):
    qs, errores, filtros_aplicados = _filtrar_prospectos(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, _ = _resumen_prospectos(qs)

    tabla_filas = [
        [p.nombre, p.correo, p.telefono, p.get_interes_display(), p.get_estado_display(),
         (p.asesor_asignado.get_full_name() or p.asesor_asignado.username) if p.asesor_asignado else 'Sin asignar',
         p.fecha.strftime('%d/%m/%Y')]
        for p in qs.order_by('-fecha')
    ]

    return _excel_generico(
        hoja_titulo='Prospectos',
        resumen_headers=['Total', 'Nuevos', 'Contactados', 'Convertidos', 'Descartados'],
        resumen_valores=[total, por_estado.get('nuevo', 0), por_estado.get('contactado', 0),
                          por_estado.get('convertido', 0), por_estado.get('descartado', 0)],
        tabla_headers=['Nombre', 'Correo', 'Teléfono', 'Interés', 'Estado', 'Asesor', 'Fecha'],
        tabla_filas=tabla_filas,
        nombre_archivo=f'reporte_prospectos_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )


# ── POSTULANTES ─────────────────────────────────────────────────────────────

def _filtrar_postulantes(request):
    from postulantes.models import Postulante
    from vacantes.models import Vacante

    qs = Postulante.objects.select_related('vacante').all()

    periodo    = request.GET.get('periodo', '').strip()
    vacante_id = request.GET.get('vacante', '').strip()
    estado     = request.GET.get('estado', '').strip()

    errores = []
    filtros_aplicados = {'periodo': None, 'vacante': None, 'estado': None}

    inicio = _rango_periodo(periodo)
    if inicio:
        qs = qs.filter(fecha__gte=inicio)
        filtros_aplicados['periodo'] = {'hoy': 'Hoy', 'semana': 'Esta semana',
                                         'mes': 'Este mes', 'año': 'Este año'}.get(periodo)

    if vacante_id:
        try:
            vacante_obj = Vacante.objects.filter(id=int(vacante_id)).first()
            qs = qs.filter(vacante_id=int(vacante_id))
            if vacante_obj:
                filtros_aplicados['vacante'] = vacante_obj.titulo
        except ValueError:
            errores.append('Vacante inválida.')

    if estado:
        if estado in dict(Postulante.ESTADO_CHOICES):
            qs = qs.filter(estado=estado)
            filtros_aplicados['estado'] = dict(Postulante.ESTADO_CHOICES)[estado]
        else:
            errores.append('Estado inválido.')

    return qs, errores, filtros_aplicados


def _resumen_postulantes(qs):
    from postulantes.models import Postulante

    total = qs.count()
    por_estado = {clave: qs.filter(estado=clave).count() for clave, _ in Postulante.ESTADO_CHOICES}

    por_vacante_map = {}
    for p in qs:
        nombre = p.vacante.titulo if p.vacante else 'Sin vacante'
        por_vacante_map[nombre] = por_vacante_map.get(nombre, 0) + 1
    por_vacante = [{'vacante': n, 'total': c} for n, c in sorted(por_vacante_map.items(), key=lambda x: -x[1])]

    return total, por_estado, por_vacante


@login_required
@solo_admin
def reporte_postulantes_ajax(request):
    qs, errores, _ = _filtrar_postulantes(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, por_vacante = _resumen_postulantes(qs)

    listado = [
        {
            'id':      p.id,
            'nombre':  p.nombre,
            'correo':  p.correo,
            'vacante': p.vacante.titulo if p.vacante else 'Sin vacante',
            'estado':  p.get_estado_display(),
            'fecha':   p.fecha.strftime('%d/%m/%Y'),
        }
        for p in qs.order_by('-fecha')[:200]
    ]

    return JsonResponse({
        'ok': True,
        'resumen': {
            'total':      total,
            'nuevo':      por_estado.get('nuevo', 0),
            'revisado':   por_estado.get('revisado', 0),
            'entrevista': por_estado.get('entrevista', 0),
            'finalista':  por_estado.get('finalista', 0),
            'contratado': por_estado.get('contratado', 0),
            'rechazado':  por_estado.get('rechazado', 0),
        },
        'por_vacante': por_vacante,
        'listado':     listado,
    })


@login_required
@solo_admin
def exportar_postulantes_pdf(request):
    qs, errores, filtros_aplicados = _filtrar_postulantes(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, _ = _resumen_postulantes(qs)
    generado = _dt.now().strftime('%d/%m/%Y %H:%M')
    filtros_texto = [v for v in filtros_aplicados.values() if v]
    subtitulo = (' | '.join(filtros_texto) if filtros_texto else 'Sin filtros aplicados') + f' — Generado: {generado}'

    tabla_filas = [
        [p.nombre, p.correo, p.vacante.titulo if p.vacante else 'Sin vacante', p.get_estado_display(),
         p.fecha.strftime('%d/%m/%Y')]
        for p in qs.order_by('-fecha')
    ]

    return _pdf_generico(
        titulo='Grupo V&amp;R Consultores — Reporte de Postulantes',
        subtitulo=subtitulo,
        resumen_headers=['Total', 'Nuevos', 'Revisados', 'Entrevista', 'Finalistas', 'Contratados', 'Rechazados'],
        resumen_valores=[total, por_estado.get('nuevo', 0), por_estado.get('revisado', 0),
                          por_estado.get('entrevista', 0), por_estado.get('finalista', 0),
                          por_estado.get('contratado', 0), por_estado.get('rechazado', 0)],
        tabla_headers=['Nombre', 'Correo', 'Vacante', 'Estado', 'Fecha'],
        tabla_filas=tabla_filas,
        col_widths=None,
        nombre_archivo=f'reporte_postulantes_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf',
    )


@login_required
@solo_admin
def exportar_postulantes_excel(request):
    qs, errores, filtros_aplicados = _filtrar_postulantes(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    total, por_estado, _ = _resumen_postulantes(qs)
    tabla_filas = [
        [p.nombre, p.correo, p.vacante.titulo if p.vacante else 'Sin vacante', p.get_estado_display(),
         p.fecha.strftime('%d/%m/%Y')]
        for p in qs.order_by('-fecha')
    ]

    return _excel_generico(
        hoja_titulo='Postulantes',
        resumen_headers=['Total', 'Nuevos', 'Revisados', 'Entrevista', 'Finalistas', 'Contratados', 'Rechazados'],
        resumen_valores=[total, por_estado.get('nuevo', 0), por_estado.get('revisado', 0),
                          por_estado.get('entrevista', 0), por_estado.get('finalista', 0),
                          por_estado.get('contratado', 0), por_estado.get('rechazado', 0)],
        tabla_headers=['Nombre', 'Correo', 'Vacante', 'Estado', 'Fecha'],
        tabla_filas=tabla_filas,
        nombre_archivo=f'reporte_postulantes_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )


@login_required
@solo_admin
def vacantes_lista_ajax(request):
    """Lista de vacantes para poblar el filtro 'Vacante' del Reporte de Postulantes."""
    from vacantes.models import Vacante
    data = [{'id': v.id, 'titulo': v.titulo} for v in Vacante.objects.order_by('titulo')]
    return JsonResponse({'ok': True, 'vacantes': data})


# ── VACANTES ─────────────────────────────────────────────────────────────
# Nota: el modelo Vacante no tiene un campo "categoría" propio; se usa el
# campo "área" como criterio de agrupación equivalente, y "modalidad" como
# dato complementario. No se crea ningún campo nuevo (fuera del alcance
# autorizado para esta fase).

@login_required
@solo_admin
def reporte_vacantes_ajax(request):
    from vacantes.models import Vacante

    Vacante.sincronizar_vencidas()
    qs = Vacante.objects.all()

    total     = qs.count()
    activas   = qs.filter(estado='activa').count()
    pausadas  = qs.filter(estado='pausada').count()
    cerradas  = qs.filter(estado='cerrada').count()
    vencidas  = qs.filter(estado='vencida').count()
    inactivas = pausadas + cerradas + vencidas

    por_area_map = {}
    for v in qs:
        por_area_map[v.area] = por_area_map.get(v.area, 0) + 1
    por_area = [{'area': a, 'total': c} for a, c in sorted(por_area_map.items(), key=lambda x: -x[1])]

    listado = [
        {
            'id':       v.id,
            'titulo':   v.titulo,
            'area':     v.area,
            'modalidad': v.get_modalidad_display(),
            'estado':   v.get_estado_display(),
            'postulantes': v.postulantes.count(),
            'creada':   v.creada.strftime('%d/%m/%Y'),
        }
        for v in qs.order_by('-creada')[:200]
    ]

    return JsonResponse({
        'ok': True,
        'resumen': {
            'total': total, 'activas': activas, 'inactivas': inactivas,
            'pausadas': pausadas, 'cerradas': cerradas, 'vencidas': vencidas,
        },
        'por_area': por_area,
        'listado':  listado,
    })


@login_required
@solo_admin
def exportar_vacantes_pdf(request):
    from vacantes.models import Vacante

    Vacante.sincronizar_vencidas()
    qs = Vacante.objects.order_by('-creada')

    total     = qs.count()
    activas   = qs.filter(estado='activa').count()
    inactivas = total - activas
    generado  = _dt.now().strftime('%d/%m/%Y %H:%M')

    tabla_filas = [
        [v.titulo, v.area, v.get_modalidad_display(), v.get_estado_display(),
         str(v.postulantes.count()), v.creada.strftime('%d/%m/%Y')]
        for v in qs
    ]

    return _pdf_generico(
        titulo='Grupo V&amp;R Consultores — Reporte de Vacantes',
        subtitulo=f'Sin filtros aplicados — Generado: {generado}',
        resumen_headers=['Total', 'Activas', 'Inactivas'],
        resumen_valores=[total, activas, inactivas],
        tabla_headers=['Título', 'Área', 'Modalidad', 'Estado', 'Postulantes', 'Creada'],
        tabla_filas=tabla_filas,
        col_widths=None,
        nombre_archivo=f'reporte_vacantes_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf',
    )


@login_required
@solo_admin
def exportar_vacantes_excel(request):
    from vacantes.models import Vacante

    Vacante.sincronizar_vencidas()
    qs = Vacante.objects.order_by('-creada')

    total     = qs.count()
    activas   = qs.filter(estado='activa').count()
    inactivas = total - activas

    tabla_filas = [
        [v.titulo, v.area, v.get_modalidad_display(), v.get_estado_display(),
         v.postulantes.count(), v.creada.strftime('%d/%m/%Y')]
        for v in qs
    ]

    return _excel_generico(
        hoja_titulo='Vacantes',
        resumen_headers=['Total', 'Activas', 'Inactivas'],
        resumen_valores=[total, activas, inactivas],
        tabla_headers=['Título', 'Área', 'Modalidad', 'Estado', 'Postulantes', 'Creada'],
        tabla_filas=tabla_filas,
        nombre_archivo=f'reporte_vacantes_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )


# ── SERVICIOS ────────────────────────────────────────────────────────────

@login_required
@solo_admin
def reporte_servicios_ajax(request):
    from servicios.models import Servicio

    qs = Servicio.objects.all()
    total    = qs.count()
    activos  = qs.filter(activo=True).count()
    inactivos = qs.filter(activo=False).count()

    por_categoria = [
        {'categoria': label, 'total': qs.filter(categoria=clave).count()}
        for clave, label in Servicio.CATEGORIA_CHOICES
    ]

    listado = [
        {
            'id':        s.id,
            'nombre':    s.nombre,
            'categoria': s.get_categoria_display(),
            'estado':    'Activo' if s.activo else 'Inactivo',
        }
        for s in qs.order_by('orden')[:200]
    ]

    return JsonResponse({
        'ok': True,
        'resumen': {'total': total, 'activos': activos, 'inactivos': inactivos},
        'por_categoria': por_categoria,
        'listado': listado,
    })


@login_required
@solo_admin
def exportar_servicios_pdf(request):
    from servicios.models import Servicio

    qs = Servicio.objects.order_by('orden')
    total     = qs.count()
    activos   = qs.filter(activo=True).count()
    inactivos = total - activos
    generado  = _dt.now().strftime('%d/%m/%Y %H:%M')

    tabla_filas = [[s.nombre, s.get_categoria_display(), 'Activo' if s.activo else 'Inactivo'] for s in qs]

    return _pdf_generico(
        titulo='Grupo V&amp;R Consultores — Reporte de Servicios',
        subtitulo=f'Sin filtros aplicados — Generado: {generado}',
        resumen_headers=['Total', 'Activos', 'Inactivos'],
        resumen_valores=[total, activos, inactivos],
        tabla_headers=['Nombre', 'Categoría', 'Estado'],
        tabla_filas=tabla_filas,
        col_widths=None,
        nombre_archivo=f'reporte_servicios_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf',
    )


@login_required
@solo_admin
def exportar_servicios_excel(request):
    from servicios.models import Servicio

    qs = Servicio.objects.order_by('orden')
    total     = qs.count()
    activos   = qs.filter(activo=True).count()
    inactivos = total - activos

    tabla_filas = [[s.nombre, s.get_categoria_display(), 'Activo' if s.activo else 'Inactivo'] for s in qs]

    return _excel_generico(
        hoja_titulo='Servicios',
        resumen_headers=['Total', 'Activos', 'Inactivos'],
        resumen_valores=[total, activos, inactivos],
        tabla_headers=['Nombre', 'Categoría', 'Estado'],
        tabla_filas=tabla_filas,
        nombre_archivo=f'reporte_servicios_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )


# ── RESUMEN GENERAL ──────────────────────────────────────────────────────

@login_required
@solo_admin
def resumen_general_ajax(request):
    from prospectos.models import Prospecto
    from postulantes.models import Postulante
    from vacantes.models import Vacante
    from servicios.models import Servicio

    Vacante.sincronizar_vencidas()

    return JsonResponse({
        'ok': True,
        'resumen': {
            'prospectos': Prospecto.objects.count(),
            'citas':      Cita.objects.count(),
            'postulantes': Postulante.objects.count(),
            'vacantes':   Vacante.objects.count(),
            'servicios':  Servicio.objects.count(),
            'usuarios_activos': User.objects.filter(is_active=True).count(),
        },
    })


@login_required
@solo_admin
def exportar_resumen_general_pdf(request):
    from prospectos.models import Prospecto
    from postulantes.models import Postulante
    from vacantes.models import Vacante
    from servicios.models import Servicio

    Vacante.sincronizar_vencidas()
    generado = _dt.now().strftime('%d/%m/%Y %H:%M')

    valores = {
        'Prospectos':          Prospecto.objects.count(),
        'Citas':               Cita.objects.count(),
        'Postulantes':         Postulante.objects.count(),
        'Vacantes':            Vacante.objects.count(),
        'Servicios':           Servicio.objects.count(),
        'Usuarios activos':    User.objects.filter(is_active=True).count(),
    }

    return _pdf_generico(
        titulo='Grupo V&amp;R Consultores — Resumen General del Sistema',
        subtitulo=f'Generado: {generado}',
        resumen_headers=[],
        resumen_valores=[],
        tabla_headers=['Módulo', 'Total'],
        tabla_filas=[[k, v] for k, v in valores.items()],
        col_widths=None,
        nombre_archivo=f'resumen_general_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf',
    )


@login_required
@solo_admin
def exportar_resumen_general_excel(request):
    from prospectos.models import Prospecto
    from postulantes.models import Postulante
    from vacantes.models import Vacante
    from servicios.models import Servicio

    Vacante.sincronizar_vencidas()

    valores = {
        'Prospectos':          Prospecto.objects.count(),
        'Citas':               Cita.objects.count(),
        'Postulantes':         Postulante.objects.count(),
        'Vacantes':            Vacante.objects.count(),
        'Servicios':           Servicio.objects.count(),
        'Usuarios activos':    User.objects.filter(is_active=True).count(),
    }

    return _excel_generico(
        hoja_titulo='Resumen General',
        resumen_headers=[],
        resumen_valores=[],
        tabla_headers=['Módulo', 'Total'],
        tabla_filas=[[k, v] for k, v in valores.items()],
        nombre_archivo=f'resumen_general_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )


# ── CITAS: exportación a Excel (paridad con el PDF ya existente) ─────────

@login_required
@solo_admin
def exportar_citas_excel(request):
    qs, errores, filtros_aplicados = _filtrar_citas(request)
    if errores:
        return JsonResponse({'ok': False, 'error': ' '.join(errores)}, status=400)

    citas = list(qs.order_by('-fecha', '-hora'))
    total       = len(citas)
    pendientes  = sum(1 for c in citas if c.estado == 'pendiente')
    confirmadas = sum(1 for c in citas if c.estado == 'confirmada')
    atendidas   = sum(1 for c in citas if c.estado == 'completada')
    canceladas  = sum(1 for c in citas if c.estado == 'cancelada')

    tabla_filas = [
        [f'{c.nombre_cliente} {c.apellidos_cliente}', c.fecha.strftime('%d/%m/%Y'), c.hora.strftime('%H:%M'),
         c.get_motivo_display(), c.get_estado_display(),
         (c.asesor.get_full_name() or c.asesor.username) if c.asesor else 'Sin asignar']
        for c in citas
    ]

    return _excel_generico(
        hoja_titulo='Citas',
        resumen_headers=['Total', 'Pendientes', 'Confirmadas', 'Atendidas', 'Canceladas'],
        resumen_valores=[total, pendientes, confirmadas, atendidas, canceladas],
        tabla_headers=['Cliente', 'Fecha', 'Hora', 'Motivo', 'Estado', 'Asesor'],
        tabla_filas=tabla_filas,
        nombre_archivo=f'reporte_citas_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )
